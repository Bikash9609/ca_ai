"""
ITC Evaluation - Single and batch invoice evaluation with working papers
"""

import json
import logging
from typing import Dict, Any, List, Optional
from datetime import datetime
from .engine import RulesEngine

logger = logging.getLogger(__name__)


class ITCEvaluator:
    """Evaluates ITC eligibility for invoices"""
    
    def __init__(self, rules_engine: RulesEngine):
        self.rules_engine = rules_engine
    
    async def evaluate_single_invoice(
        self,
        invoice: Dict[str, Any],
        gstr2b_data: Optional[Dict[str, Any]] = None,
        registration_data: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Evaluate a single invoice for ITC eligibility
        
        Returns:
            {
                "invoice_number": str,
                "vendor_gstin": str,
                "invoice_date": str,
                "taxable_value": float,
                "tax_amount": float,
                "eligible": bool,
                "eligible_amount": float,
                "blocked_amount": float,
                "rules_applied": List[Dict],
                "explanation": str,
                "recommendation": str
            }
        """
        result = await self.rules_engine.evaluate_invoice(
            invoice,
            gstr2b_data,
            registration_data
        )
        
        # Add recommendation
        if result["blocked_amount"] > 0:
            recommendation = (
                f"ITC of ₹{result['blocked_amount']:.2f} is blocked. "
                f"Review the rules applied and take necessary action."
            )
        elif result["eligible"]:
            recommendation = "ITC is eligible. Can be claimed in GSTR-3B."
        else:
            recommendation = "ITC is not eligible. Do not claim in GSTR-3B."
        
        return {
            **result,
            "invoice_date": invoice.get("invoice_date"),
            "taxable_value": invoice.get("taxable_value", 0.0),
            "recommendation": recommendation
        }
    
    async def evaluate_batch_invoices(
        self,
        invoices: List[Dict[str, Any]],
        gstr2b_data: Optional[Dict[str, Any]] = None,
        registration_data: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Evaluate multiple invoices"""
        batch_result = await self.rules_engine.evaluate_batch(
            invoices,
            gstr2b_data,
            registration_data
        )
        
        return batch_result
    
    def generate_working_paper(
        self,
        evaluation_results: Dict[str, Any],
        period: str,
        client_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Generate working paper from evaluation results
        
        Returns:
            {
                "period": str,
                "client_name": str,
                "summary": Dict,
                "detailed_breakdown": List[Dict],
                "rule_citations": List[str],
                "generated_at": str
            }
        """
        results = evaluation_results.get("results", [])
        
        # Summary calculations
        total_invoices = len(results)
        eligible_invoices = sum(1 for r in results if r["eligible"])
        blocked_invoices = sum(1 for r in results if r["blocked_amount"] > 0)
        
        total_tax = sum(r["total_tax"] for r in results)
        total_eligible = sum(r["eligible_amount"] for r in results)
        total_blocked = sum(r["blocked_amount"] for r in results)
        
        # Collect all rule citations
        rule_citations = set()
        for result in results:
            for rule in result.get("rules_applied", []):
                citation = rule.get("citation")
                if citation:
                    rule_citations.add(citation)
        
        # Detailed breakdown
        detailed_breakdown = []
        for result in results:
            breakdown = {
                "invoice_number": result.get("invoice_number"),
                "vendor_gstin": result.get("vendor_gstin"),
                "taxable_value": result.get("taxable_value", 0.0),
                "total_tax": result.get("total_tax", 0.0),
                "eligible_amount": result.get("eligible_amount", 0.0),
                "blocked_amount": result.get("blocked_amount", 0.0),
                "rules_applied": [
                    {
                        "rule_id": r.get("rule_id"),
                        "rule_name": r.get("rule_name"),
                        "citation": r.get("citation"),
                        "action": r.get("action", {}).get("action")
                    }
                    for r in result.get("rules_applied", [])
                ],
                "status": "eligible" if result["eligible"] else "blocked"
            }
            detailed_breakdown.append(breakdown)
        
        return {
            "period": period,
            "client_name": client_name,
            "summary": {
                "total_invoices": total_invoices,
                "eligible_invoices": eligible_invoices,
                "blocked_invoices": blocked_invoices,
                "total_tax_amount": total_tax,
                "total_eligible_itc": total_eligible,
                "total_blocked_itc": total_blocked,
                "eligibility_percentage": (total_eligible / total_tax * 100) if total_tax > 0 else 0
            },
            "detailed_breakdown": detailed_breakdown,
            "rule_citations": sorted(list(rule_citations)),
            "generated_at": datetime.utcnow().isoformat()
        }
    
    def export_working_paper_json(
        self,
        working_paper: Dict[str, Any],
        file_path: str
    ) -> None:
        """Export working paper as JSON"""
        with open(file_path, "w") as f:
            json.dump(working_paper, f, indent=2)
        logger.info(f"Exported working paper to {file_path}")
    
    def export_working_paper_excel(
        self,
        working_paper: Dict[str, Any],
        file_path: str
    ) -> None:
        """Export working paper as Excel (requires openpyxl)"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "ITC Working Paper"
            
            # Header
            ws["A1"] = "ITC Working Paper"
            ws["A1"].font = Font(bold=True, size=14)
            ws.merge_cells("A1:F1")
            
            # Summary section
            row = 3
            ws[f"A{row}"] = "Period:"
            ws[f"B{row}"] = working_paper.get("period", "")
            row += 1
            
            summary = working_paper.get("summary", {})
            ws[f"A{row}"] = "Total Invoices:"
            ws[f"B{row}"] = summary.get("total_invoices", 0)
            row += 1
            
            ws[f"A{row}"] = "Total Eligible ITC:"
            ws[f"B{row}"] = summary.get("total_eligible_itc", 0.0)
            row += 1
            
            ws[f"A{row}"] = "Total Blocked ITC:"
            ws[f"B{row}"] = summary.get("total_blocked_itc", 0.0)
            row += 2
            
            # Detailed breakdown headers
            headers = ["Invoice Number", "Vendor GSTIN", "Taxable Value", 
                      "Total Tax", "Eligible Amount", "Blocked Amount", "Status"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            row += 1
            
            # Data rows
            for item in working_paper.get("detailed_breakdown", []):
                ws.cell(row=row, column=1).value = item.get("invoice_number", "")
                ws.cell(row=row, column=2).value = item.get("vendor_gstin", "")
                ws.cell(row=row, column=3).value = item.get("taxable_value", 0.0)
                ws.cell(row=row, column=4).value = item.get("total_tax", 0.0)
                ws.cell(row=row, column=5).value = item.get("eligible_amount", 0.0)
                ws.cell(row=row, column=6).value = item.get("blocked_amount", 0.0)
                ws.cell(row=row, column=7).value = item.get("status", "")
                row += 1
            
            # Rule citations sheet
            ws2 = wb.create_sheet("Rule Citations")
            ws2["A1"] = "Rule Citations"
            ws2["A1"].font = Font(bold=True)
            
            for idx, citation in enumerate(working_paper.get("rule_citations", []), start=2):
                ws2.cell(row=idx, column=1).value = citation
            
            wb.save(file_path)
            logger.info(f"Exported working paper to Excel: {file_path}")
        except ImportError:
            logger.warning("openpyxl not installed. Cannot export to Excel. Install with: uv pip install openpyxl")
            raise
